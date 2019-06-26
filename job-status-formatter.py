###################################################
#  File:    excel.py
#  Author:  nathan
#  Date:    27/09/2018
#  Brief:   a python program to automatically format the JOB REPORTS spreadsheet
###################################################

from openpyxl.styles import PatternFill, Font, Alignment
import openpyxl
import os
import shutil
import datetime

print(" N4tH4N's")
print("     __      ___.")
print("    |__| ____\\_ |__       ")                                 
print("    |  |/  _ \\| __ \\      ")
print("    |  (  <_> ) \\_\\ \\     ")
print("/\\__|  |\\____/|___  /     ")
print("\\______|          \\/             ")
print("          __          __                                   ")
print("  _______/  |______ _/  |_ __ __  ______                   ")
print(" /  ___/\\   __\\__  \\\\   __\\  |  \\/  ___/                   ")
print(" \\___ \\  |  |  / __ \\|  | |  |  /\\___ \\                    ")
print("/____  > |__| (____  /__| |____//____  >                   ")
print("     \\/            \\/                \\/                    ")
print("  _____                            __    __                ")
print("_/ ____\\___________  _____ _____ _/  |__/  |_  ___________ ")
print("\\   __\\/  _ \\_  __ \\/     \\\\__  \\\\   __\\   __\\/ __ \\_  __ \\ ")
print(" |  | (  <_> )  | \\/  Y Y  \\/ __ \\|  |  |  | \\  ___/|  | \\/ ")
print(" |__|  \\____/|__|  |__|_|  (____  /__|  |__|  \\___  >__|   ")
print("                         \\/     \\/                \\/       ")
print("version 1.1 - BUG FIX, not removing the correct rows.")
print("version 1.2 - BUG FIX, added a 'try' and 'except' to the sorting function.")
print("                       This will catch errors if there are rows with blank due dates.")
print("version 1.3 - UPDATE, 'JIGS' client code will now be deleted.")
print("version 1.4 - UPDATE, Now highlights jobs which are [ LASER CUT ONLY ].")
print("version 1.5 - BUG FIX, row values getting mixed up with the LASER CUT ONLY jobs.")
print("                       rewrote the entire function from scratch, so far so good... more testing needed.")
print("version 1.6 - UPDATE, The SUB column is now included when searching for [ LASER CUT ONLY ] parts")
print("version 1.7 - UPDATE, The paint column has been replaced with STOPA & CLEAN")
print("                      the STOPA column will be positioned after the 3030 column")
print("version 1.8 - UPDATE, The columns it checks for LASER CUT ONLY need to be updated after adding the STOPA column")
print("version 1.9 - BUG FIX, put '90 XPNT' back into the correct_order array, with no 53 BSAW")
print("                       this was causing issues.")
print("version 2.0 - UPDATE, Now highlights jobs that will be [ PACK AT PRESS ]")
print("version 2.1 - BUG FIX, the addition of SAW column in the schedule was breaking")
print("                       the PACK AT PRESS function.")
print("version 2.2 - UPDATE, pack at press colour updated.")
print("version 2.3 - UPDATE, Refining the PACK AT PRESS function.")
print("")

###############################
#### EXCEL FILE TO WORK ON ####
file_name = 'book1.xlsx'
###############################

######################################################
####  testing if a spreadsheet to work on exists  ####
######################################################
file_exists = os.path.isfile(file_name)
if not file_exists:
    print("")
    print("")
    print("")
    print("There has been an,               ._. ")
    print("  __________________  ___________| | ")
    print("_/ __ \\_  __ \\_  __ \\/  _ \\_  __ \\ | ")
    print("\\  ___/|  | \\/|  | \\(  <_> )  | \\/\\| ")
    print(" \\___  >__|   |__|   \\____/|__|   __ ")
    print("     \\/                           \\/ ")
    print("")
    print("COULD NOT FiND ANY FiLE CALLED 'Book1.xlsx'")
    print("")
    input("Press any ENTER to exit...")
    exit()  # exiting the program

###############################  used to style the cells
red_background_colour = PatternFill("solid", fgColor="ffc7ce")  # This will fill the cell RED
yellow_background_colour = PatternFill("solid", fgColor="ffff00")  # This will fill the cell YELLOW
bussy_blue_background_colour = PatternFill("solid", fgColor="d4f4f9") # blue for bustech
all_clocked_colour = PatternFill("solid", fgColor="C6EFCE")  # all clocked colour is green
laser_cut_only_colour = PatternFill("solid", fgColor="32e3ff")  # jobs that are laser cut only
pack_at_press_colour = PatternFill("solid", fgColor="006ec0")  # jobs that are laser cut only

white_font_colour = Font(color="ffffff")
###############################

CURRENT_DIRECTORY = os.getcwd()
now = datetime.datetime.now()
todays_date_string = now.strftime("%Y-%m-%d")  # creating a date string format: YYYY-mm-dd
old_file_name = todays_date_string + " " + file_name
old_files_folder = os.path.join(CURRENT_DIRECTORY, '_original_files')

workbook = openpyxl.load_workbook(file_name)
worksheet = workbook.active

print('Loading all the data from the spreadsheet into memory...')

all_rows = []  # the whole spreadsheet will be loaded into this array
headings = []

for row_counter, row in enumerate(worksheet, 1):
    if row_counter > 1:  # ignoring the first row
        current_row = []
        for column_counter, cell in enumerate(row, 1):
            if column_counter != 2:  # the 'Order Date' column is not needed
                current_row.append(cell.value)
        all_rows.append(current_row)
    else:
        for column_counter, cell in enumerate(row, 1):
            if column_counter != 2:
                headings.append(cell.value)

total_number_of_rows = ( len(all_rows) + 1 )
print( "There are {} rows to process.".format(total_number_of_rows))

# sorting alphabetically by the 'Client Code' column
# print("Sorting all the data by the 'Client Code' column...")
# all_rows = sorted(all_rows, key=lambda client: client[2])
print("Sorting all the data by 'Due Date' then by 'Client Code' then by 'Job No'...")
try:
    all_rows.sort(key = lambda l: (l[1], l[2], l[0]))
except TypeError:
    print("")
    print("There has been an,               ._. ")
    print("  __________________  ___________| | ")
    print("_/ __ \\_  __ \\_  __ \\/  _ \\_  __ \\ | ")
    print("\\  ___/|  | \\/|  | \\(  <_> )  | \\/\\| ")
    print(" \\___  >__|   |__|   \\____/|__|   __ ")
    print("     \\/                           \\/ ")
    print("Is there a row in your spreadsheet that has a blank due date??")
    print("")
    input("Press any ENTER to exit...")
    exit()  # exiting the program


print("Removing 'DRAW', 'SCHEDULE', 'TESTCLIENT', 'GCI-NON-PRODUCTIVE-TIME' from the spreadsheet")
print("and clearing the contents of the 'Job Status' column...")
index_numbers_to_delete = []
for counter, line in enumerate(all_rows):
    line[5] = ""  # clearing contents of the 'Job Status' column
    if line[2] == 'DRAW' or line[2] == 'SCHEDULE' or line[2] == 'TESTCLIENT' or line[2] == 'GCI-NON-PRODUCTIVE-TIME':  # testing the 'Client Code' column
        index_numbers_to_delete.append(counter)
    if line[4] == "All":  # deleting all rows that have 'All' in the 'Desp' column
        index_numbers_to_delete.append(counter)

# deleting the arrays that have 'DRAW', 'SCHEDULE' or 'TESTCLIENT' in Client Code column
index_numbers_to_delete.reverse()  # reversing the array avoids any index errors when deleting array items
for x in index_numbers_to_delete:
    del all_rows[x]

total_number_of_rows = ( len(all_rows) + 1 )
print("There are now {} rows left to process.".format(total_number_of_rows))

#######################################################
####  starting to order the departments correctly  ####
#######################################################

correct_order = [  # this is the correct order that 'G' wants
    "1 PROG",
    "4 3030",
    "44 STPCLN",
    "56 LISMAC",
    "6 ROTO",
    "53 BSAW",
    "51 FOLD",
    "58 GMAC",
    "67 PEMS",
    "7 TIG",
    "52 MIG",
    "90 XPNT",  ##  taking this our for now, I think it's been replace with STOPA & CLEAN  (PUT BACK IN 10/10/2019, NO SAW IN THE SCHEDULE BROKE A LOT OF THINGS)
    "36 SANDBL",
    "21 PC",
    "Sub"
]

####  !important - we use this array to re-organise the columns
correct_index_order = [0,1,2,3,4,5,6]  # the columns from 0 to 6 do not need to move
####
index_dictionary = {}  # creating a dictionary out of the 'headings'
for counter, item in enumerate(headings):
    index_dictionary.update({item : counter})

for department in correct_order:  #
    if department in index_dictionary:  # checking to see if the department is in the dictionary
        # print("{} is at index: {}".format(department, index_dictionary[department]))
        correct_index_order.append(index_dictionary[department])
        if department == "51 FOLD":
            print('Found the FOLD column')
            FOLD_COLUMN = ( len(correct_index_order) - 1 )
    else:
        print("could not find {}".format(department))

print("Correct index order,")
print(correct_index_order)

## rearranging the headings array
re_ordered_headings = []
for index_number in correct_index_order:
    re_ordered_headings.append(headings[index_number])

################################################
####  moving the original spreadsheet file  ####
################################################

# moving the original spreadsheet to the '_original_files' folder with the date appended to the file name
print("Moving the original spreadsheet to the '_original_files' folder...")
# shutil.move(file_name, os.path.join(old_files_folder, old_file_name))

########################################
####  creating the new spreadsheet  ####
########################################

print("Creating the new spreadsheet...")
new_work_book = openpyxl.Workbook()
new_sheet = new_work_book.active

# filling in the first row with the headings
print("Filling in the headings row...")
for counter, title in enumerate(re_ordered_headings, 1):
    # renaming the column headings
    if title == "1 PROG":
        title = "PROG"
    elif title == "4 3030":
        title = 3030
    elif title == "44 STPCLN":
        title = "STOPA"
    elif title == "56 LISMAC":
        title = "LIS"
    elif title == "6 ROTO":
        title = "ROTO"
    elif title == "53 BSAW":
        title = "SAW"
    elif title == "51 FOLD":
        title = "FOLD"
    elif title == "58 GMAC":
        title = "GMAC"
    elif title == "67 PEMS":
        title = "PEMS"
    elif title == "7 TIG":
        title = "TIG"
    elif title == "52 MIG":
        title = "MIG"
    elif title == "90 XPNT":  ##  PAINT HAS BEEN REMOVED FROM THE SPREADSHEET FOR NOW... (PUT BACK IN 10/01/2019 SEE ^^ FOR EXPLANATION)
        title = "PNT"
    elif title == "36 SANDBL":
        title = "SBL"
    elif title == "21 PC":
        title = "PC"
    new_sheet.cell(row=1, column=counter).value = title
    new_sheet.cell(row=1, column=counter).font = Font(bold=True) # heading will be bold
    if counter >= 5:  # column 5 is the 'Desp' column
        # every heading after and including the 'Desp' heading will have center text
        new_sheet.cell(row=1, column=counter).alignment = Alignment(horizontal='center')
    if counter <=2:  # column 1 and 2 have center text
        new_sheet.cell(row=1, column=counter).alignment = Alignment(horizontal='center')
    if title == "Sub":
        new_sheet.cell(row=1, column=counter).fill = red_background_colour

# filling in all the data from the second row on
print("Filling in the rest of the data...")
all_rows_reordered = []  # storing the re-ordered rows to use later for the styling of cells
for row_counter, line in enumerate(all_rows, 2):
    re_ordered_rows = []  # temp array for the re-ordered row
    for index_number in correct_index_order:  # looping through the correct_index_order array
        re_ordered_rows.append(line[index_number])  # adding the contents to the re_ordered_row
    for column_counter, item in enumerate(re_ordered_rows, 1):
        if column_counter == 2:  # column 2 is Due Date, have to format the 'datetime' object back to a string object
            date_formatted = item.strftime('%d/%m/%Y')  # format == dd/mm/YYYY
            new_sheet.cell(row=row_counter, column=column_counter).value = date_formatted
        else:
            new_sheet.cell(row=row_counter, column=column_counter).value = item
        if column_counter >= 5 or column_counter <= 2:  # column 5 is the 'Desp' column
            # every heading after and including the 'Desp' heading will have center text
            new_sheet.cell(row=row_counter, column=column_counter).alignment = Alignment(horizontal='center')
    all_rows_reordered.append(re_ordered_rows)

#############################################
####  starting to style the spreadsheet  ####
#############################################

print("Highlighting the 'Client Code' column (red) for:")
print("EXTERNAL-RECUT, RECUT-INTERNAL, MISSEDPROCESS, REWORK-INTERNAL, INTERNAL, ADDITION_2_CURRENT_JOB...")
highlight_client_code_array = []
highlight_bustech_array = []
for row_counter, client in enumerate(all_rows, 2):
    if client[2] == "EXTERNAL-RECUT":
        highlight_client_code_array.append(row_counter)
    elif client[2] == "EXTERNAL-REWORK":
        highlight_client_code_array.append(row_counter)
    elif client[2] == "RECUT-INTERNAL":
        highlight_client_code_array.append(row_counter)
    elif client[2] == "MISSEDPROCESS":
        highlight_client_code_array.append(row_counter)
    elif client[2] == "REWORK-INTERNAL":
        highlight_client_code_array.append(row_counter)
    elif client[2] == "INTERNAL":
        highlight_client_code_array.append(row_counter)
    elif client[2] == "ADDITION_2_CURRENT_JOB":
        highlight_client_code_array.append(row_counter)
    elif client[2] == "BUSTECH":  # BUSTECH gets highlighted a blue colour
        highlight_bustech_array.append(row_counter)

for x in highlight_client_code_array:
    column = 'C'
    row = x
    cell = new_sheet[(column + str(row))]
    cell.fill = red_background_colour  # G requested that the client code cell be highlighted RED

for x in highlight_bustech_array:  # highlighting all the client code cells that have 'BUSTECH'
    column = 'C'
    row = x
    cell = new_sheet[(column + str(row))]
    cell.fill = bussy_blue_background_colour

print("Finding all the cells that have 'Part' in the Desp column...")
highlight_desp_column_array = []
for row_counter, value in enumerate(all_rows_reordered, 2):
    if value[4] == "Part":
        highlight_desp_column_array.append(row_counter)

print("Highlighting all the cells in the Desp column that has 'Part'...")
for x in highlight_desp_column_array:
    column = 'E'
    row = x
    cell = new_sheet[(column + str(row))]
    cell.fill = red_background_colour

print("Finding all the cells that have 'Sub' in the Sub column...")
highlight_sub_column_array = []
for row_counter, value in enumerate(all_rows_reordered, 2):
    if value[20] == 'Sub':
        highlight_sub_column_array.append(row_counter)

print("Highlighting all the cells in the Sub column that has 'Sub'...")
for x in highlight_sub_column_array:
    column = 'U'
    row = x
    cell = new_sheet[(column + str(row))]
    cell.fill = red_background_colour

# if row is all clocked and the customer is in this array, the row will be deleted
clients_to_delete_if_row_all_clocked = [
    "RECUT-INTERNAL", "MISSEDPROCESS", "OBSOLETE-PROCESS",
    "REWORK-INTERNAL", "INTERNAL", "ADDITION_2_CURRENT_JOB", "JIGS"
]
rows_to_delete = []

print("Finding which cells contain an unfinished process...")
highlight_process_array = []
all_clocked_array = []
for row_counter, row in enumerate(all_rows_reordered, 2):
    all_clocked = False
    for column_counter, column in enumerate(row, 1):
        if column_counter >= 8 and column is not None and column != "Sub":
            if isinstance(column, int):  # testing if the cell is an integer
                continue  # if the cell is an integer it will continue to the next loop
            column_split = column.split()
            if len(column_split) > 1:
                left_value = column_split[0]
                right_value = column_split[2]
                if int(left_value) > int(right_value):  # testing if the left value is greater than the right value e.g. 4 | 1
                    column_letter = chr(column_counter + 64)  # see http://www.asciitable.com/ (A = 1, Dec for A is 65)
                    highlight_process_array.append([column_letter, row_counter])
                    all_clocked = True
    if all_clocked == False:
        all_clocked_array.append(row_counter)
        # if all the processes have been done and if the client is in the 'clients_to_delete_if_row_all_clocked'
        # the row number is appended to rows to delete
        if row[2] in clients_to_delete_if_row_all_clocked:
            rows_to_delete.append(row_counter)

print("Entering 'All clocked' on jobs that have no unfinished processes...")
for x in all_clocked_array:
    new_sheet.cell(row=x, column=6).value = "All clocked"
    new_sheet.cell(row=x, column=6).alignment = Alignment(horizontal='center')   
    new_sheet.cell(row=x, column=6).fill = all_clocked_colour

print("Highlighting all the cells with an unfinished process...")
for x in highlight_process_array:
    # the 'highlight_process_array' will have values like [column letter, row number]
    cell_coordinate = x[0] + str(x[1])  # e.g. 'N3'
    cell = new_sheet[cell_coordinate]
    cell.fill = yellow_background_colour

print("Finding the rows which are LASER only...")
laser_only_rows = []
for row_counter, row in enumerate(all_rows_reordered, 2):
    row_is_laser_only = True
    for column_counter, column in enumerate(row, 1):
        if column_counter >= 11:  #  only looking at columns past STOPA & CLEAN
            if column != None:
                row_is_laser_only = False
    if row_is_laser_only:
        laser_only_rows.append(row_counter)

print("Highlighting all the rows which are LASER only...")
for x in laser_only_rows:
    cell_coordinate = "D{0}".format(x)
    cell = new_sheet[cell_coordinate]
    text = cell.value
    print("{0} - {1}".format(x, text))
    cell.fill = laser_cut_only_colour
    cell.font = white_font_colour
    cell.value = "{0}  [ LASER CUT ONLY ]".format(text)

print("Finding the rows which will be PACK AT PRESS...")
pack_at_press_rows = []
for row_counter, row in enumerate(all_rows_reordered, 2):
    # checking if there is a value in the FOLD column
    # if there is a value, it will proceed to check if this row
    # is pack at press
    if row[FOLD_COLUMN] != None:
        row_is_pack_at_press = True

        # testing if 'SAW' column is in the schedule
        # if it is the fold column will be along by one
        if "53 BSAW" in re_ordered_headings:
            # if SAW is in the schedule the fold column will be 14
            fold_column_number = 14
        else:
            # if SAW is not in the schedule the fold column will be 13
            fold_column_number = 13
            
        for column_counter, column in enumerate(row, 1):
            if column_counter > fold_column_number:  # only looking at columns past FOLD
                if column != None:
                    row_is_pack_at_press = False
        if row_is_pack_at_press:
            pack_at_press_rows.append(row_counter)
    else:
        # skipping over the row if there is no
        # value in the FOLD column
        continue

print("Highlighting all the rows which are PACK AT PRESS...")
for x in pack_at_press_rows:
    # checking if the row number is already in the LASER CUT ONLY row
    if x in laser_only_rows:
        continue  # if it is, we will skip it
    cell_coordinate = "D{0}".format(x)
    cell = new_sheet[cell_coordinate]
    text = cell.value
    print("{0} - {1}".format(x, text))
    cell.fill = pack_at_press_colour
    cell.font = white_font_colour
    cell.value = "{0}  [ PACK AT PRESS ]".format(text)

print("Deleting rows from the spreadsheet if the 'Job Status' is all clocked,")
print("    and the customer is in the 'clients_to_delete_if_row_all_clocked'")
rows_to_delete.reverse()  #  important to reverse the list so it deletes rows from the bottom
for row_number in rows_to_delete:
    new_sheet.delete_rows(row_number, 1)

################################
####  resizing the columns  ####
################################

print("Resizing the column widths...")
column_width_array = [ ["A", 7], ["B", 12], ["C", 25], ["D", 80], ["E", 7], 
                    ["F", 14], ["G", 6.5], ["H", 9.5], ["I", 9.5], ["J", 7], ["K", 9.5],
                    ["L", 7], ["M", 7], ["N", 8], ["O", 7], ["P", 7], ["Q", 7],
                    ["R", 7], ["S", 7], ["T", 7], ["U", 6]]
for y in column_width_array:
    new_sheet.column_dimensions[y[0]].width = y[1]

######################################
####  saving the new spreadsheet  ####
######################################

print("Saving the new spreadsheet...")
new_file_name = todays_date_string + " " + "new_file.xlsx"  # saving the new file with todays date added to the start of the file
new_work_book.save(new_file_name)
print("File saved.")

print("")
